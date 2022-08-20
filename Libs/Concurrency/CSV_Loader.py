import os.path
from functools import lru_cache
from datetime import datetime

import openpyxl
import pandas as pd
from PyQt5 import QtCore

from Libs.globals import *
from Libs.Utils import config

logger = exception_handler.getFutureLogger(__name__)


class SymbLoader(QtCore.QObject):
    """To load instrument.csv"""
    loaded = QtCore.pyqtSignal(object)
    completed = QtCore.pyqtSignal()
    load_path = settings.DATA_FILES["symbols_mapping_csv"]

    @lru_cache(maxsize=None)
    def load_symbol_names(self):
        instruments_file_path = settings.DATA_FILES.get('INSTRUMENTS_CSV')
        symbol_mapping_df = pd.read_csv(self.load_path)
        symbol_names = symbol_mapping_df["NFO_CODE"].values.tolist()
        if not os.path.exists(instruments_file_path):
            logger.info("Downloading instruments file...")
            config.download_instruments_file()
        instruments_df = pd.read_csv(instruments_file_path)
        instruments_df['days2expire'] = (pd.to_datetime(instruments_df.expiry).dt.date - datetime.today().date()).dt.days
        instruments_df = instruments_df[(instruments_df['days2expire'] < 120) |
                                        (
                                                (instruments_df['days2expire'].isna()) &
                                                (instruments_df['instrument_type'] == "EQ")
                                        )]
        self.loaded.emit((symbol_names, instruments_df))
        self.completed.emit()


class StrategySavedLoader(QtCore.QObject):
    data_row = QtCore.pyqtSignal(tuple)  # List[Union[int, str]], strategy_name
    strategies_loaded = QtCore.pyqtSignal(dict)

    def __init__(self):
        super().__init__()
        self.workbook = None
        self._saved_data_dict = {}

    def load_workbook_data(self):
        try:
            self.workbook = openpyxl.load_workbook(settings.DATA_FILES['tradexcb_excel_file'])
            for row_index, row in enumerate(self.workbook['Sheet1'].rows, 0):
                if row_index == 0:
                    continue
                row_data_list = [col.value for col in row]
                if row_data_list[-1] not in self._saved_data_dict:
                    self._saved_data_dict[row_data_list[-1]] = []
                self._saved_data_dict[row_data_list[-1]].append(row_data_list[:-1])
            self.strategies_loaded.emit(self._saved_data_dict)
        except FileNotFoundError:
            logger.warning("No saved strategies found. Creating new file.")
            # create excel file
            self.workbook = openpyxl.Workbook()
            self.workbook.create_sheet("Sheet1")
            self.workbook.save(settings.DATA_FILES['tradexcb_excel_file'])
            logger.info("New file for saving strategies created.")
            self.strategies_loaded.emit(self._saved_data_dict)
