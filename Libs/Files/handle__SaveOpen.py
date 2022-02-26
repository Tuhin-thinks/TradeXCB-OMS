import re
import openpyxl
from PyQt5 import QtWidgets, QtCore

import Libs
from Libs.Files.TradingSymbolMapping import StrategiesColumn
from Libs.globals import *

logger = exception_handler.getFutureLogger(__name__)


def open_file(parent, caption, directory, filter_str):
    open_path, _ = QtWidgets.QFileDialog().getOpenFileName(parent=parent, caption=caption, directory=directory,
                                                           filter=filter_str)
    return open_path


def save_file(parent, caption, directory, filter_str):
    file_dialog = QtWidgets.QFileDialog()
    options = QtWidgets.QFileDialog.Options()
    options |= file_dialog.DontConfirmOverwrite
    save_path_url, selected_filter = file_dialog.getSaveFileUrl(parent=parent, caption=caption,
                                                                directory=QtCore.QUrl.fromLocalFile(directory),
                                                                filter=filter_str,
                                                                options=options)
    file_path = save_path_url.toLocalFile()
    search_obj = re.search(r"\(\*\.(\w+)\)", selected_filter)
    if search_obj:
        extension = search_obj.groups()[0]
        if not file_path.endswith(f".{extension}"):
            file_path = f"{file_path}.{extension}"
    return file_path


def delete_excel_sheet(sheet):
    max_rows = sheet.max_column - 1
    max_cols = sheet.max_row - 1
    sheet.delete_cols(1, max_cols)
    sheet.delete_rows(1, max_rows)


def export_as_excel(parent: 'Libs.api_home.ApiHome'):
    excel_path = settings.DATA_FILES["Delta_plus_Algo_File"]
    complete_save_data = []
    strategy_name_index_mapping = {}
    strategies_to_save = set()
    with Time__Profiler.ProfilerContext("saving File"):
        row_index = 0
        for strategy_table in parent.strategy_tables:
            list_dict = strategy_table.save()
            if list_dict:
                strategy_name = strategy_table.name()
                complete_save_data.extend(list_dict)
                strategies_to_save.add(strategy_name)
                for row_ in range(row_index, row_index + len(list_dict)):
                    strategy_name_index_mapping[row_] = strategy_name
                row_index += len(list_dict)
        field_names = StrategiesColumn.delta_plus_algo_columns + ['strategy_name']
        try:
            workbook = openpyxl.load_workbook(excel_path)

            main_sheet = workbook["Sheet1"]
            # delete sheet and write the header
            delete_excel_sheet(main_sheet)
            for col_index, key in enumerate(field_names, 1):
                main_sheet.cell(row=1, column=col_index).value = key

            # --------------- for loop based in 1-based indexing in openpyxl -----------------
            # ------------------------- write header row for main sheet --------------
            # ---------------- iterate on all data to save them in respective excel --------------
            for col_index in range(1, len(field_names) + 1):
                key = field_names[col_index - 1]
                mm_row_index = 2
                for row_index in range(2, len(complete_save_data) + 2):
                    cell_value = complete_save_data[row_index - 2].get(key)  # 0-based-indexing for python use.
                    if cell_value is None:
                        continue
                    main_sheet.cell(row=mm_row_index, column=col_index).value = cell_value
                    mm_row_index += 1

            workbook.save(excel_path)
            parent.show_status("strategies saved successfully")
            return 0
        except Exception as ex:
            parent.show_status("cannot save strategies")
            logger.critical(f"Failed to save strategies: {ex.__str__()}", exc_info=True)
            return -1
